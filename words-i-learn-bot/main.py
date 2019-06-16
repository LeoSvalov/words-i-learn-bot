import telebot
from telebot import types
from openpyxl import load_workbook
from random import choice
import time

# class for 1 word. It contains word, explanation(in russian) and the link with precise details(optional)


class Slot:
    def __init__(self, word, meaning, link):
        self.word = word
        self.meaning = meaning
        self.link = link


wb = load_workbook('./words_I_learn.xlsx')
sheet = wb.active
s = list()
b = sheet.max_row
b = str(b)
first_cell = "A2"
last_cell = "C" + b
counter = 0
for cellObj in sheet[first_cell:last_cell]:
    i = 0
    for cell in cellObj:
        if i == 0:
            word = cell.value
            i = i + 1
        elif i == 1:
            meaning = cell.value
            i = i + 1
        else:
            link = cell.value
            i = i + 1
            if link:
                counter = counter + 1
    temp = Slot(word, meaning, link)
    s.append(temp)

bot = telebot.TeleBot("711239282:AAFb5sqOwL4tkVVczQoqZzYGIZyDM0dzHQk")

# start - Beginning of the bot
@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id, "Hey, *dude*! Let's start learning some english! ", parse_mode='Markdown')

# help - Description of all commands
@bot.message_handler(commands=['help'])
def help_message(message):
    bot.send_message(message.chat.id, '*All commands:* \n' +
                     '/word - It gives u a random word.\n' +
                     '/revise - It gives u random 10 words.\n' +
                     "/quiz - It gives u 1 minute to translate 10 random words and shows the answers.\n" +
                     '_Hope u will expand your vocabulary and improve yourself!_', parse_mode='Markdown')

# word - It takes a random words from dictionary
@bot.message_handler(commands=['word'])
def word_message(message):
    random_word = choice(s)
    if random_word.link:
        bot.send_message(message.chat.id, text="*word* - "
                                               + random_word.word + "\n" + "*meaning* - "
                                               + random_word.meaning + "\n"
                                               + "*More details* [here]("
                                               + random_word.link + ")" + ".", parse_mode='Markdown')
    else:
        bot.send_message(message.chat.id, text="*word* - "
                                               + random_word.word + "\n"
                                               + "*meaning* - "
                                               + random_word.meaning, parse_mode='Markdown')

# revise - It takes 10 random words from dictionary
@bot.message_handler(commands=['revise'])
def revise_message(message):
    x = list()
    while len(x) < 10:
        q = choice(s)
        if not x.__contains__(q):
            x.append(q)
    text = ""
    for i in range(0, 10):
        a_word = x[i]
        if a_word.link:
           text = text + str(i+1) + ".\n" + "*word* - " + a_word.word + "\n" + "*meaning* - " + a_word.meaning + "\n" + "*More details* [here](" + a_word.link + ")" + "." + "\n"
        else:
            text = text + str(i+1) + ".\n" + "*word* - " + a_word.word + "\n" + "*meaning* - " + a_word.meaning + ".\n"
    bot.send_message(message.chat.id, text, parse_mode='Markdown')

# quiz - It takes 10 random words from dictionary and shows answers after 1 minute.
@bot.message_handler(commands=['quiz'])
def quiz_message(message):
    x = list()
    while len(x) < 10:
        q = choice(s)
        if not x.__contains__(q):
            x.append(q)
    text = ""

    for i in range(0, 10):
        a_word = x[i]
        if a_word.link:
            text = text + str(i+1) + ".\n" + "*word* - " + a_word.word + "\n"
        else:
            text = text + str(i+1) + ".\n" + "*word* - " + a_word.word + "\n"
    text = text + '\n' + '*Answers will be after 1 minute.*\n'
    bot.send_message(message.chat.id, text, parse_mode='Markdown')
    time.sleep(60)
    text = ""
    for i in range(0, 10):
        a_word = x[i]
        if a_word.link:
            text = text + str(i+1) + ".\n" + "*word* - " + a_word.word + "\n" + "*meaning* - " +\
                   a_word.meaning + "\n" + "*More details* [here](" + a_word.link + ")" + "." + "\n"
        else:
            text = text + str(i+1) + ".\n" + "*word* - " + a_word.word + "\n" + "*meaning* - " + a_word.meaning + ".\n"
    bot.send_message(message.chat.id, text, parse_mode='Markdown')
    time.sleep(0.5)
    markup = types.ReplyKeyboardMarkup(row_width=2)
    item1 = types.KeyboardButton('1-2')
    item2 = types.KeyboardButton('3-4')
    item3 = types.KeyboardButton('5-6')
    item4 = types.KeyboardButton('7-8')
    item5 = types.KeyboardButton('9-10')
    markup.row(item1, item2, item3, item4, item5)
    bot.send_message(message.chat.id, "*How many did u answer correctly?*",  parse_mode='Markdown', reply_markup=markup)
    markup = types.ReplyKeyboardRemove(selective=True)
    time.sleep(4)
    bot.send_sticker(message.chat.id, 'CAADAQADvwgAAr-MkAQ34mkMgYWhPAI', reply_markup=markup)


@bot.message_handler(content_types=['text'])
def send_message(message):
    if message.text == 'Hi' or  message.text == 'hi' or  message.text == 'hello' or  message.text == 'Hello':
        bot.send_message(message.chat.id,'Hey, dude!')
    elif message.text == 'how are you?' or message.text == 'How are you?' or message.text == 'how are u?' or message.text == 'How are u?' or message.text == "What's up?" or message.text == "what's up?":
        bot.send_message(message.chat.id,'I am completely fine. I am only a telegram bot, so have never had any troubles. Hope u too!')
        bot.send_sticker(message.chat.id, 'CAADAgADxwMAAsSraAvTQCJCCeKWBwI')
    elif message.text != '1-2' and message.text != '3-4' and message.text != '5-6' and message.text != '7-8' and message.text != '9-10':
        bot.send_message(message.chat.id,"Sorry, I can't understand you. (•_•)")

bot.polling()